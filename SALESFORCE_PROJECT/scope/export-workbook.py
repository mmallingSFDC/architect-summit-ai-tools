#!/usr/bin/env python3
"""
Generate scoping deliverables Excel workbook
"""
import json
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def auto_size_columns(ws):
    """Auto-size columns based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column_letter].width = adjusted_width

def add_csv_sheet(wb, csv_path, sheet_name):
    """Add a sheet from CSV file"""
    if not csv_path.exists():
        return None

    ws = wb.create_sheet(sheet_name)

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Header row formatting
                if row_idx == 1:
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Auto-size columns
    auto_size_columns(ws)

    # Freeze header row
    ws.freeze_panes = 'A2'

    return ws

def main():
    project_dir = Path('.')
    data_dir = project_dir / 'data'
    csv_dir = data_dir / 'csv'
    outputs_dir = project_dir / 'outputs'

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Add sheets in order
    sheets = [
        (csv_dir / '01-epics.csv', 'Epics'),
        (csv_dir / '02-estimates.csv', 'Estimates'),
        (csv_dir / '03-roadmap-phases.csv', 'Roadmap Phases'),
        (csv_dir / '04-roles.csv', 'Roles & Skills'),
    ]

    # Check for user stories (story-level project)
    user_stories_csv = csv_dir / '01-user-stories.csv'
    if user_stories_csv.exists():
        sheets.insert(1, (user_stories_csv, 'User Stories'))

    sheets_added = []
    for csv_path, sheet_name in sheets:
        if csv_path.exists():
            add_csv_sheet(wb, csv_path, sheet_name)
            sheets_added.append(sheet_name)
            print(f"[OK] Added sheet: {sheet_name}")
        else:
            print(f"[SKIP] Not found: {sheet_name}")

    # Save workbook
    output_path = outputs_dir / 'scoping-deliverables.xlsx'
    wb.save(output_path)

    print(f"\n[SUCCESS] Excel workbook generated: {output_path}")
    print(f"   Sheets: {', '.join(sheets_added)}")

    return output_path

if __name__ == '__main__':
    main()
